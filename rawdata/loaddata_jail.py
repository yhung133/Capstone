from openpyxl import load_workbook
import json

data = []
json_path = "jailbook.json"

wb = load_workbook('BOOKRJTF.xlsx')
sheetnames = wb.get_sheet_names()
count = 1
for i in range(0,len(sheetnames)):
    sheet = wb.get_sheet_by_name(name=sheetnames[i])
    for row in sheet.iter_rows(row_offset=1):
        tempdict = {}
        tempdict["model"] = "racialjustice.jailbook"
        tempdict["pk"] = count
        count += 1
        fields = ["book_date","book_time","brith_date","age","gender","race","number","direction",
                  "street","suffix","city","state","zipcode","statute","description","court_date",
                  "release_date","bond"]
        tempdict["fields"] = {}
        for j in range(0,len(fields)):
            value = row[j].internal_value
            if j == 0 or j == 2 or j == 15 or j == 16:
                try:
                    date = value.split('/')
                    for k in range(0,3):
                        if len(date[k]) == 1:
                            date[k] = '0' + date[k]
                    value = date[2] + '-' + date[0] + '-' + date[1]
                except:
                    value = '0000-00-00'
            if j == 3:
                try:
                    value = int(row[0].split('/')[2]) - int(row[2].split('/')[2])
                except:
                    value = 0
            tempdict["fields"][fields[j]] = value
        data.append(tempdict)
        
print len(date)

with io.open(json_path, 'w', encoding = 'utf-8') as jsonfile:
    json_str = json.dumps(data, indent=4, sort_keys=True, separators=(',',':'), ensure_ascii=False)
    if isinstance(json_str, str):
        try:
            json_str = json_str.decode('utf-8')
        except:
            json_str = json_str.decode('latin-1')
    jsonfile.write(json_str)




